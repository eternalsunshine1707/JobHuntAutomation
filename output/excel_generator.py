"""
Excel Generator - Creates a professionally formatted job listings spreadsheet.
"""
import os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import OUTPUT_DIR

# EST timezone (UTC-5), EDT (UTC-4)
EST = timezone(timedelta(hours=-4))  # EDT during daylight saving


def generate_excel(jobs: list[dict]) -> str:
    """Generate a formatted Excel file with job listings. Returns filepath."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    now_est = datetime.now(EST)
    timestamp = now_est.strftime("%Y%m%d_%I%M%p_EST")
    filename = f"job_listings_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Job Listings"

    # Styles
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(name="Arial", size=10, color="006100")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    yellow_font = Font(name="Arial", size=10, color="9C6500")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(name="Arial", size=10, color="9C0006")
    alt_fill = PatternFill("solid", fgColor="F2F7FB")
    border = Border(
        left=Side("thin", "B0B0B0"), right=Side("thin", "B0B0B0"),
        top=Side("thin", "B0B0B0"), bottom=Side("thin", "B0B0B0"),
    )

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Job Listings — {now_est.strftime('%B %d, %Y at %I:%M %p EST')}"
    ws["A1"].font = title_font
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Data Engineer & Data Analyst | All US | F-1 OPT — Visa Sponsorship Prioritized"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")

    # Headers
    headers = ["#", "Company", "Job Title", "Location", "Source", "Posted", "Applicants", "Match Score", "Visa Sponsorship", "Apply Link"]
    widths = [5, 25, 40, 22, 12, 16, 14, 13, 18, 55]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 28

    # Data rows
    for idx, job in enumerate(jobs):
        row = idx + 5
        values = [
            idx + 1,
            job.get("company", ""),
            job.get("title", ""),
            job.get("location", ""),
            job.get("source", ""),
            job.get("posted_time", "Unknown"),
            str(job.get("applicants", "Unknown")),
            job.get("match_score", "N/A"),
            job.get("visa_sponsorship", "Unknown"),
            job.get("apply_link", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if idx % 2 == 1:
                cell.fill = alt_fill

        # Hyperlink for apply link
        link_cell = ws.cell(row=row, column=10)
        link_cell.font = link_font
        link_url = job.get("apply_link", "")
        if link_url and link_url.startswith("http"):
            try:
                link_cell.hyperlink = link_url
            except Exception:
                pass

        # Color-code visa column
        visa_cell = ws.cell(row=row, column=9)
        visa_val = str(job.get("visa_sponsorship", "")).lower()
        if "yes" in visa_val:
            visa_cell.fill = green_fill
            visa_cell.font = green_font
        elif "no" in visa_val:
            visa_cell.fill = red_fill
            visa_cell.font = red_font
        else:
            visa_cell.fill = yellow_fill
            visa_cell.font = yellow_font

        ws.row_dimensions[row].height = 30

    # Freeze header
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:J{4 + len(jobs)}"

    wb.save(filepath)
    print(f"  [Excel] Saved {len(jobs)} jobs to {filepath}")
    return filepath
